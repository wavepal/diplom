from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.urls import reverse
from index.models import Choices, User, Form, Responses, Answer, Questions, RegionMedCenter, MedCenterGroup, CITY_CHOICES
import json
import csv
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from django.db.models import Avg, Q
from dateutil.relativedelta import relativedelta
from django.core.paginator import Paginator, EmptyPage
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.utils.http import quote as urlquote

@csrf_exempt
def delete_selected_responses(request):
    if request.method == "POST":
        data = json.loads(request.body)
        response_ids = data.get("response_ids", [])

        if response_ids:
            Responses.objects.filter(id__in=response_ids).delete()
            return JsonResponse({"success": True})

    return JsonResponse({"success": False}, status=400)

def calculate_average_scores(responses, form):
    average_scores = {}
    
    # Группируем ответы по медцентрам
    for response in responses:
        med_center = response.responder_med
        if med_center not in average_scores:
            average_scores[med_center] = {
                'scores': {},
                'counts': {},
                'total_score': 0
            }
        
        # Обрабатываем каждый ответ
        for answer in response.response.filter(is_skipped=False):
            question = answer.answer_to
            if question.question_type == "range slider":
                if question.id not in average_scores[med_center]['scores']:
                    average_scores[med_center]['scores'][question.id] = 0
                    average_scores[med_center]['counts'][question.id] = 0
                
                try:
                    value = float(answer.answer)
                    # Просто суммируем исходные значения
                    average_scores[med_center]['scores'][question.id] += value
                    average_scores[med_center]['counts'][question.id] += 1
                except (ValueError, TypeError):
                    continue

    # Вычисляем средние значения
    for med_center in average_scores:
        total_valid_questions = 0
        total_percentage = 0
        
        for question_id in average_scores[med_center]['scores']:
            count = average_scores[med_center]['counts'][question_id]
            if count > 0:
                # Вычисляем среднее значение и процент от максимального
                question = form.questions.get(id=question_id)
                raw_average = average_scores[med_center]['scores'][question_id] / count
                average_scores[med_center]['scores'][question_id] = raw_average
                if question.max_value != 0:
                    percentage = (raw_average / question.max_value) * 100
                else:
                    percentage = 0
                total_percentage += percentage
                total_valid_questions += 1
        
        # Вычисляем общий средний процент
        if total_valid_questions > 0:
            average_scores[med_center]['total_score'] = round(total_percentage / total_valid_questions, 2)
        else:
            average_scores[med_center]['total_score'] = None

    return average_scores

def responses(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))

    if not (request.user.is_admin() or request.user.is_trainer() or request.user.is_manager()):
        return HttpResponseRedirect(reverse("403"))

    formInfo = get_object_or_404(Form.objects.prefetch_related(
        'questions',
        'questions__choices'
    ), code=code)

    # Получаем параметры фильтрации
    selected_city = request.GET.get('cities')
    selected_gender = request.GET.get('gender')
    age_min = request.GET.get('age_min')
    age_max = request.GET.get('age_max')
    selected_med_region = request.GET.get('med_region')
    selected_med_center = request.GET.get('med_center')
    selected_med_group = request.GET.get('med_group')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Базовый запрос с минимально необходимыми данными для отображения страницы
    all_responses = Responses.objects.filter(response_to=formInfo).select_related(
        'responder'
    )
    
    # Применяем фильтры с использованием Q objects для оптимизации
    filters = Q()
    if selected_city:
        filters &= Q(responder_city=selected_city)
    if selected_gender:
        filters &= Q(responder_gender=selected_gender)
    if age_min and age_min.strip():
        try:
            filters &= Q(responder_age__gte=int(age_min))
        except (ValueError, TypeError):
            pass
    if age_max and age_max.strip():
        try:
            filters &= Q(responder_age__lte=int(age_max))
        except (ValueError, TypeError):
            pass
    if selected_med_group:
        med_centers = RegionMedCenter.objects.filter(group_id=selected_med_group).values_list('med_center', flat=True)
        filters &= Q(responder_med__in=med_centers)
    if selected_med_region:
        med_centers = RegionMedCenter.objects.filter(region=selected_med_region).values_list('med_center', flat=True)
        filters &= Q(responder_med__in=med_centers)
    if selected_med_center and selected_med_center.strip():
        # Ensure exact match for med_center
        filters &= Q(responder_med__exact=selected_med_center)
    if date_from and date_from.strip():
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            filters &= Q(createdAt__gte=date_from_obj)
        except (ValueError, TypeError):
            pass
    if date_to and date_to.strip():
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            filters &= Q(createdAt__lt=date_to_obj)
        except (ValueError, TypeError):
            pass

    all_responses = all_responses.filter(filters)

    # Process questions and answers
    responsesSummary, choiceAnswered, choices_dict = process_questions_and_answers(formInfo)
    
    # Calculate filtered response summary
    filteredResponsesSummary = get_filtered_response_summary(choiceAnswered, all_responses, formInfo)
    
    # Get range slider data
    range_slider_data = get_range_slider_data(formInfo, all_responses)

    # Подготавливаем базовый контекст с минимально необходимыми данными
    context = {
        "form": formInfo,
        "responses": all_responses,
        "responsesSummary": responsesSummary,
        "choiceAnswered": choiceAnswered,
        "choices_dict": choices_dict,
        "filteredResponsesSummary": filteredResponsesSummary,
        "range_slider_data": range_slider_data,
        "city_choices": CITY_CHOICES,
        'med_centers': RegionMedCenter.objects.values_list('med_center', flat=True).distinct(),
        'med_center_groups': MedCenterGroup.objects.all().order_by('name'),
        'active_forms': Form.objects.filter(is_active=True),
    }

    return render(request, "index/form/analytics.html", context)

def get_filtered_responses(formInfo, age_min, age_max, selected_gender, selected_cities, selected_med_center=None):
    all_responses = Responses.objects.filter(response_to=formInfo)
    
    # Используем сохраненные статические данные вместо связей
    if age_min and age_min.strip():
        try:
            all_responses = all_responses.filter(responder_age__gte=int(age_min))
        except (ValueError, TypeError):
            pass
    if age_max and age_max.strip():
        try:
            all_responses = all_responses.filter(responder_age__lte=int(age_max))
        except (ValueError, TypeError):
            pass
    if selected_gender:
        all_responses = all_responses.filter(responder_gender=selected_gender)
    if selected_cities and selected_cities != ['']:
        all_responses = all_responses.filter(responder_city__in=selected_cities)
    if selected_med_center and selected_med_center.strip():
        all_responses = all_responses.filter(responder_med__exact=selected_med_center)
    
    return all_responses

def process_questions_and_answers(formInfo):
    responsesSummary = []
    choiceAnswered = {}
    choices_dict = {}

    for question in formInfo.questions.exclude(question_type="title"):
        answers = Answer.objects.filter(answer_to=question.id, is_skipped=False)

        if question.question_type in ["multiple choice", "checkbox"]:
            choiceAnswered[question.id] = choiceAnswered.get(question.id, {})
            for answer in answers:
                try:
                    choice = answer.answer_to.choices.get(id=answer.answer)
                    choices_dict[answer.answer] = choice.choice
                    unique_choice_key = f"{choice.choice}"
                    choiceAnswered[question.id][unique_choice_key] = choiceAnswered[question.id].get(unique_choice_key, 0) + 1
                except Choices.DoesNotExist:
                    continue

        responsesSummary.append({"question": question, "answers": answers})
    
    return responsesSummary, choiceAnswered, choices_dict

def get_range_slider_data(formInfo, all_responses):
    data = {}
    end_date = datetime.now()
    start_date = end_date - timedelta(days=365)
    current_month = end_date.replace(day=1)
    months = [(current_month - relativedelta(months=i)).strftime('%B %Y') for i in range(12)]
    months.reverse()

    questions = formInfo.questions.filter(question_type="range slider").exclude(question_type="title")
    
    # Get all responses for the form, ignoring date filters
    all_form_responses = Responses.objects.filter(response_to=formInfo)
    
    # Extract med_center filter from all_responses if it exists
    med_centers_filter = None
    if all_responses.count() > 0:
        # Check if med_center filter is applied (by comparing counts)
        if all_responses.values('responder_med').distinct().count() < all_form_responses.values('responder_med').distinct().count():
            med_centers_filter = list(all_responses.values_list('responder_med', flat=True).distinct())
    
    for question in questions:
        try:
            # Base query for responses_for_question
            responses_query = Answer.objects.filter(
                answer_to=question,
                response__in=all_form_responses,
                response__createdAt__gte=start_date,
                is_skipped=False
            )
            
            # Apply med_center filter if it exists
            if med_centers_filter:
                responses_query = responses_query.filter(response__responder_med__in=med_centers_filter)
                
            responses_for_question = responses_query.select_related('response')
            
            # Get unique centers
            centers = responses_for_question.values_list('response__responder_med', flat=True).distinct()
            monthly_averages_by_center = {center: [0] * 12 for center in centers if center is not None}

            # Calculate monthly averages using raw SQL to avoid OperationalError
            monthly_query = Answer.objects.filter(
                answer_to=question,
                response__in=all_form_responses,
                response__createdAt__gte=start_date,
                is_skipped=False
            )
            
            # Apply med_center filter if it exists
            if med_centers_filter:
                monthly_query = monthly_query.filter(response__responder_med__in=med_centers_filter)
                
            monthly_data = monthly_query.extra(
                select={'month': "strftime('%%m', createdAt)"}
            ).values(
                'response__responder_med',
                'month'
            ).annotate(
                avg_value=Avg('answer')
            )

            for item in monthly_data:
                center = item['response__responder_med']
                if center is not None and center in monthly_averages_by_center:
                    try:
                        month = int(item['month'])
                        month_idx = (end_date.month - month) % 12
                        if 0 <= month_idx < 12:
                            avg_value = item['avg_value']
                            if avg_value is not None:
                                monthly_averages_by_center[center][11 - month_idx] = round(float(avg_value), 2)
                    except (ValueError, TypeError, IndexError) as e:
                        continue

            data[question.id] = {
                'months': months,
                'averages_by_center': monthly_averages_by_center,
                'max_value': question.max_value
            }
        except Exception as e:
            print(f"Error in get_range_slider_data: {e}")
            continue
    
    return data

def get_response_answers(all_responses, formInfo):
    data = {}
    questions = formInfo.questions.prefetch_related('choices').exclude(question_type="title")
    
    for response in all_responses:
        data[response.id] = {}
        answers = response.response.select_related('answer_to').prefetch_related('answer_to__choices')
        
        for answer in answers:
            question = answer.answer_to
            if not answer.is_skipped:
                if question.question_type == "checkbox":
                    selected_choices = []
                    choice_ids = answer.answer.split(',') if answer.answer else []
                    choices = question.choices.filter(id__in=choice_ids)
                    selected_choices = [choice.choice for choice in choices]
                    data[response.id][question.id] = selected_choices if selected_choices else "N/A"
                elif question.question_type == "multiple choice":
                    try:
                        choice = question.choices.get(id=answer.answer)
                        data[response.id][question.id] = choice.choice
                    except Choices.DoesNotExist:
                        data[response.id][question.id] = "N/A"
                else:
                    data[response.id][question.id] = answer.answer
            else:
                data[response.id][question.id] = "N/A"
                    
    return data

def get_filtered_response_summary(choiceAnswered, all_responses, formInfo):
    # Создаем словарь для хранения результатов
    filteredResponsesSummary = {}
    
    print(f"DEBUG: Starting get_filtered_response_summary with {all_responses.count()} responses")
    
    # Инициализируем счетчики для каждого вопроса и варианта ответа
    for question in formInfo.questions.exclude(question_type="title"):
        if question.question_type in ["multiple choice", "checkbox"]:
            filteredResponsesSummary[question.id] = {}
            for choice in question.choices.all():
                filteredResponsesSummary[question.id][choice.choice] = 0
    
    print(f"DEBUG: Initialized filteredResponsesSummary with {len(filteredResponsesSummary)} questions")
    
    # Для каждого вопроса и ответа напрямую
    for response in all_responses:
        for question in formInfo.questions.filter(question_type__in=["multiple choice", "checkbox"]):
            answers = Answer.objects.filter(
                response=response,
                answer_to=question,
                is_skipped=False
            )
            
            for answer in answers:
                if question.question_type == "checkbox":
                    try:
                        choice = question.choices.get(id=answer.answer)
                        filteredResponsesSummary[question.id][choice.choice] += 1
                        print(f"DEBUG: Added count for question {question.id}, choice {choice.choice}")
                    except Choices.DoesNotExist:
                        continue
                elif question.question_type == "multiple choice":
                    try:
                        choice = question.choices.get(id=answer.answer)
                        filteredResponsesSummary[question.id][choice.choice] += 1
                        print(f"DEBUG: Added count for question {question.id}, choice {choice.choice}")
                    except Choices.DoesNotExist:
                        continue
    
    # Print the final data structure for debugging
    for question_id, choices in filteredResponsesSummary.items():
        print(f"DEBUG: Question {question_id} has {len(choices)} choices with data:")
        for choice, count in choices.items():
            if count > 0:
                print(f"  - {choice}: {count}")
    
    return filteredResponsesSummary

def get_user_city_dict(all_responses):
    # Больше не нужно делать дополнительные запросы к UserCity
    user_city_dict = {}
    for response in all_responses:
        user_city_dict[response.id] = response.responder_city
    return user_city_dict

def get_average_data(formInfo, all_responses):
    average_data = {}
    questions = formInfo.questions.prefetch_related('choices').exclude(question_type="title")

    for question in questions:
        if question.question_type == "range slider":
            question_average_data = {}
            responders = all_responses.values_list('responder_med', flat=True).distinct()

            for responder_med in responders:
                selected_responses = all_responses.filter(responder_med=responder_med)
                range_answers = Answer.objects.filter(
                    response__in=selected_responses,
                    answer_to=question.id,
                    is_skipped=False
                )

                if range_answers.exists():
                    average_value = range_answers.aggregate(Avg('answer'))['answer__avg']
                    question_average_data[responder_med] = average_value if average_value is not None else 0
                else:
                    question_average_data[responder_med] = 0

            average_data[question.id] = question_average_data
    
    return average_data

def get_med_center_stats(formInfo, all_responses):
    med_center_stats = {}
    med_centers = RegionMedCenter.objects.all()

    # Инициализируем статистику используя ID медцентра
    for med_center in med_centers:
        med_center_stats[med_center.med_center] = {
            'total_responses': 0,
            'questions': {},
        }
        for question in formInfo.questions.filter(is_negative=True):
            med_center_stats[med_center.med_center]['questions'][question.id] = 0

    # Подсчитываем статистику
    for response in all_responses:
        med_center = response.responder_med
        if med_center in med_center_stats:
            med_center_stats[med_center]['total_responses'] += 1
            
            # Используем response вместо answers
            for answer in response.response.all():
                if answer.answer_to.is_negative and answer.answer == 'Да':
                    med_center_stats[med_center]['questions'][answer.answer_to.id] += 1

    return med_center_stats


def export_responses_to_excel(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))

    formInfo = get_object_or_404(Form, code=code)

    if not request.user.is_superuser and not request.user.is_staff:
        return HttpResponseRedirect(reverse("403"))

    # Получаем состояния чекбоксов и настройки отображения
    export_as_percentage = request.GET.get('export_as_percentage') == 'true'
    export_total_as_percentage = request.GET.get('export_total_as_percentage') == 'true'
    hidden_columns = request.GET.get('export_hidden_columns', '').split(',')
    sort_column = request.GET.get('sort_column')
    sort_direction = request.GET.get('sort_direction', 'asc')

    # Получаем параметры фильтрации из сессии и GET-параметров
    filter_params = request.session.get('filter_params', {})
    
    # Получаем параметры фильтрации
    selected_city = request.GET.get('cities', filter_params.get('cities'))
    selected_gender = request.GET.get('gender', filter_params.get('gender'))
    age_min = request.GET.get('age_min', filter_params.get('age_min'))
    age_max = request.GET.get('age_max', filter_params.get('age_max'))
    selected_med_region = request.GET.get('med_region', filter_params.get('med_region'))
    selected_med_center = request.GET.get('med_center', filter_params.get('med_center'))
    selected_med_group = request.GET.get('med_group', filter_params.get('med_group'))
    date_from = request.GET.get('date_from', filter_params.get('date_from'))
    date_to = request.GET.get('date_to', filter_params.get('date_to'))
    
    # Базовый запрос
    all_responses = Responses.objects.filter(response_to=formInfo)
    
    # Применяем фильтры с использованием Q objects для оптимизации
    filters = Q()
    if selected_city:
        filters &= Q(responder_city=selected_city)
    if selected_gender:
        filters &= Q(responder_gender=selected_gender)
    if age_min and age_min.strip():
        try:
            filters &= Q(responder_age__gte=int(age_min))
        except (ValueError, TypeError):
            pass
    if age_max and age_max.strip():
        try:
            filters &= Q(responder_age__lte=int(age_max))
        except (ValueError, TypeError):
            pass
    if selected_med_group:
        med_centers = RegionMedCenter.objects.filter(group_id=selected_med_group).values_list('med_center', flat=True)
        filters &= Q(responder_med__in=med_centers)
    if selected_med_region:
        med_centers = RegionMedCenter.objects.filter(region=selected_med_region).values_list('med_center', flat=True)
        filters &= Q(responder_med__in=med_centers)
    if selected_med_center and selected_med_center.strip():
        filters &= Q(responder_med__exact=selected_med_center)
    if date_from and date_from.strip():
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            filters &= Q(createdAt__gte=date_from_obj)
        except (ValueError, TypeError):
            pass
    if date_to and date_to.strip():
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            filters &= Q(createdAt__lt=date_to_obj)
        except (ValueError, TypeError):
            pass

    # Применяем фильтры по выбранным вариантам ответов
    for key, value in request.GET.items():
        if key.startswith('question-') and value:
            question_id = key.replace('question-', '')
            try:
                question = Questions.objects.get(id=question_id)
                if question.question_type in ["multiple choice", "checkbox"]:
                    choice_ids = request.GET.getlist(key)
                    if choice_ids:
                        answers = Answer.objects.filter(
                            answer_to=question,
                            answer__in=choice_ids
                        )
                        response_ids = answers.values_list('response_id', flat=True)
                        filters &= Q(id__in=response_ids)
            except Questions.DoesNotExist:
                pass

    all_responses = all_responses.filter(filters)

    # Применяем сортировку
    if sort_column:
        if sort_column == 'user':
            sort_field = 'responder_username'
        elif sort_column == 'age':
            sort_field = 'responder_age'
        elif sort_column == 'gender':
            sort_field = 'responder_gender'
        elif sort_column == 'city':
            sort_field = 'responder_city'
        elif sort_column == 'med':
            sort_field = 'responder_med'
        elif sort_column == 'created-at':
            sort_field = 'createdAt'
        else:
            sort_field = 'createdAt'

        if sort_direction == 'desc':
            sort_field = f'-{sort_field}'
        all_responses = all_responses.order_by(sort_field)

    response_answers = {}
    for response in all_responses:
        response_answers[response.id] = {}
        for question in formInfo.questions.exclude(question_type="title"):
            if question.question_type == "checkbox":
                answers = Answer.objects.filter(response=response, answer_to=question.id, is_skipped=False)
                selected_choices = []
                for answer in answers:
                    try:
                        choice = answer.answer_to.choices.get(id=answer.answer)
                        selected_choices.append(choice.choice)
                    except Choices.DoesNotExist:
                        continue
                response_answers[response.id][question.id] = selected_choices if selected_choices else "N/A"
            else:
                answer = Answer.objects.filter(response=response, answer_to=question.id, is_skipped=False).first()
                if answer:
                    if question.question_type == "multiple choice":
                        try:
                            choice = question.choices.get(id=answer.answer)
                            response_answers[response.id][question.id] = choice.choice
                        except Choices.DoesNotExist:
                            response_answers[response.id][question.id] = "N/A"
                    elif question.question_type == "range slider":
                        try:
                            raw_value = float(answer.answer)
                        except ValueError:
                            raw_value = 0.0
                        max_range_value = question.max_value
                        if export_as_percentage:
                            percentage = (raw_value / max_range_value) * 100 if max_range_value != 0 else 0
                            response_answers[response.id][question.id] = f"{percentage:.2f}%"
                        else:
                            response_answers[response.id][question.id] = raw_value
                    elif question.question_type in ["short", "paragraph"]:
                        response_answers[response.id][question.id] = answer.answer if answer.answer.strip() else "N/A"
                    else:
                        response_answers[response.id][question.id] = answer.answer
                else:
                    response_answers[response.id][question.id] = "N/A"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Responses"

    # Определяем заголовки с учетом скрытых столбцов
    base_headers = {
        "user": "User",
        "age": "Age",
        "gender": "Gender",
        "city": "City",
        "med": "Medical Center",
        "created-at": "Submission Date"
    }

    headers = []
    for key, header in base_headers.items():
        if key not in hidden_columns:
            headers.append(header)

    question_headers = {question.question: question.id for question in formInfo.questions.exclude(question_type="title")}
    for question, question_id in question_headers.items():
        if f"question-{question_id}" not in hidden_columns:
            headers.append(question)

    # Стили для заголовков
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Записываем заголовки и применяем стили
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Записываем данные
    for row_num, response in enumerate(all_responses, 2):
        col = 1
        if "user" not in hidden_columns:
            ws.cell(row=row_num, column=col, value=response.responder_username if response.responder_username else "Anonymous")
            col += 1
        if "age" not in hidden_columns:
            ws.cell(row=row_num, column=col, value=response.responder_age if response.responder_age else "N/A")
            col += 1
        if "gender" not in hidden_columns:
            ws.cell(row=row_num, column=col, value=response.responder_gender if response.responder_gender else "N/A")
            col += 1
        if "city" not in hidden_columns:
            ws.cell(row=row_num, column=col, value=response.responder_city if response.responder_city else "N/A")
            col += 1
        if "med" not in hidden_columns:
            ws.cell(row=row_num, column=col, value=response.responder_med if response.responder_med else "N/A")
            col += 1
        if "created-at" not in hidden_columns:
            ws.cell(row=row_num, column=col, value=response.createdAt.strftime("%d.%m.%Y %H:%M"))
            col += 1

        for question, question_id in question_headers.items():
            if f"question-{question_id}" not in hidden_columns:
                answer = response_answers[response.id].get(question_id, "N/A")
                if isinstance(answer, list):  # For checkbox type questions
                    answer = ", ".join(answer)
                ws.cell(row=row_num, column=col, value=answer)
                col += 1

    # Автоматическая настройка ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
        ws.column_dimensions[column_letter].width = adjusted_width

    # Применяем форматирование к ячейкам с данными
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=responses_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

def export_combined_excel(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))

    formInfo = get_object_or_404(Form, code=code)

    if not request.user.is_superuser and not request.user.is_staff:
        return HttpResponseRedirect(reverse("403"))

    visible_columns = request.GET.get('visible_columns', '').split(',')
    visible_med_centers = request.GET.get('visible_med_centers', '').split(',')
    export_as_percentage = request.GET.get('export_as_percentage') == 'true'
    export_total_as_percentage = request.GET.get('export_total_as_percentage') == 'true'

    # Получаем параметры фильтрации из сессии
    filter_params = request.session.get('filter_params', {})
    
    # Применяем те же фильтры, что и в представлении responses
    all_responses = Responses.objects.filter(response_to=formInfo)
    
    if filter_params.get('cities'):
        all_responses = all_responses.filter(responder__city_info__city=filter_params['cities'])
    
    if filter_params.get('gender'):
        all_responses = all_responses.filter(responder__gender_info__gender=filter_params['gender'])
    
    if filter_params.get('age_min'):
        all_responses = all_responses.filter(responder_age__gte=int(filter_params['age_min']))
    
    if filter_params.get('age_max'):
        all_responses = all_responses.filter(responder_age__lte=int(filter_params['age_max']))

    if filter_params.get('med_region'):
        med_centers = RegionMedCenter.objects.filter(
            region=filter_params['med_region']
        ).values_list('med_center', flat=True)
        all_responses = all_responses.filter(responder_med__in=med_centers)

    if filter_params.get('med_center'):
        all_responses = all_responses.filter(responder_med=filter_params['med_center'])

    # Применяем фильтры по выбранным вариантам ответов
    for question_id, selected_choices in filter_params.get('choices', {}).items():
        answers = Answer.objects.filter(
            response__in=all_responses,
            answer_to_id=question_id,
            answer__in=selected_choices
        )
        response_ids = answers.values_list('response_id', flat=True)
        all_responses = all_responses.filter(id__in=response_ids)

    average_scores = calculate_average_scores(all_responses, formInfo)
    average_scores_sorted = sorted(
        [item for item in average_scores.items() if item[1]['total_score'] is not None],
        key=lambda x: x[1]['total_score'],
        reverse=True
    )
    med_center_stats = get_med_center_stats(formInfo, all_responses)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined Data"

    # Стили для заголовков
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Заголовки для средних оценок
    headers = ["Медицинский центр"]
    for question in formInfo.questions.filter(question_type="range slider"):
        if f'{question.id}' in visible_columns:
            headers.append(f"{question.question}")
    
    # Добавляем заголовки для количества ответов
    for question in formInfo.questions.filter(question_type__in=["short", "paragraph"]):
        if f'answ-{question.id}' in visible_columns:
            headers.append(f"Кол-во ответов: {question.question}")
    
    if 'main-value' in visible_columns:
        headers.append("Общая оценка")
    if 'total-responses' in visible_columns:
        headers.append("Общее количество ответов")

    # Записываем заголовки и применяем стили
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)

    # Заполняем данные
    row_num = 2
    for med_center, scores in average_scores_sorted:
        if med_center not in visible_med_centers:
            continue
        
        row = [med_center]
        
        # Добавляем значения range slider
        for question in formInfo.questions.filter(question_type="range slider"):
            if f'{question.id}' in visible_columns:
                raw_value = scores['scores'].get(question.id, "N/A")
                if raw_value != "N/A" and export_as_percentage:
                    max_value = question.max_value
                    percentage = (raw_value / max_value) * 100
                    row.append(f"{percentage:.2f}%")
                else:
                    row.append(raw_value)
        
        # Добавляем количество ответов для каждого вопроса
        for question in formInfo.questions.filter(question_type__in=["short", "paragraph"]):
            if f'answ-{question.id}' in visible_columns:
                answers_count = med_center_stats[med_center]['questions'].get(question.id, 0)
                row.append(answers_count)
        
        # Добавляем общую оценку
        if 'main-value' in visible_columns:
            total_score = scores['total_score']
            if export_total_as_percentage:
                total_score = f"{total_score:.2f}%"
            row.append(total_score)
        
        if 'total-responses' in visible_columns:
            row.append(med_center_stats[med_center]['total_responses'])
        
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        row_num += 1

    # Настраиваем ширину столбцов
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Combined_Data_{formInfo.code}.xlsx'
    wb.save(response)

    return response

def retrieve_checkbox_choices(response, question):
    checkbox_answers = []

    answers = Answer.objects.filter(answer_to=question, response=response, is_skipped=False)
    for answer in answers:
        selected_choice_ids = answer.answer.split(',')  
        selected_choices = Choices.objects.filter(pk__in=selected_choice_ids)
        checkbox_answers.append([choice.choice for choice in selected_choices])

    return checkbox_answers

def exportcsv(request,code):
    formInfo = Form.objects.filter(code = code)
    formInfo = formInfo[0]
    responses=Responses.objects.filter(response_to = formInfo)
    questions = formInfo.questions.all()


    http_response = HttpResponse()
    http_response['Content-Type'] = 'text/csv'
    http_response['Content-Disposition'] = f'attachment; filename={formInfo.title}.csv'
    writer = csv.writer(http_response)
    header = ['Response Code', 'Responder', 'Responder Email','Responder_ip']

    for question in questions:
        header.append(question.question)

    writer.writerow(header)

    for response in responses:
        response_data = [
        response.response_code,
        response.responder_username if response.responder else 'Anonymous',
        response.responder_email if response.responder_email else '',
        response.responder_ip if response.responder_ip else ''
    ]
        for question in questions:
            answer = Answer.objects.filter(answer_to=question, response=response, is_skipped=False).first()


            if  question.question_type not in ['multiple choice','checkbox']:
                response_data.append(answer.answer if answer else '')
            elif question.question_type == "multiple choice":
                response_data.append(answer.answer_to.choices.get(id = answer.answer).choice if answer else '')
            elif question.question_type == "checkbox":
                if answer and question.question_type == 'checkbox':
                    checkbox_choices = retrieve_checkbox_choices(response,answer.answer_to)
                    response_data.append(checkbox_choices)

        print(response_data)
        writer.writerow(response_data)

    return http_response

def response(request, code, response_code):
    formInfo = Form.objects.filter(code=code)
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else:
        formInfo = formInfo[0]
    # if not formInfo.allow_view_score:
    #     if not request.user.is_staff:
    #         return HttpResponseRedirect(reverse("403"))

    total_score = 0
    score = 0
    responseInfo = Responses.objects.filter(response_code=response_code)
    if responseInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else:
        responseInfo = responseInfo[0]

    # if formInfo.is_quiz:
    #     for question in formInfo.questions.all():
    #         if question.question_type == "multiple choice":
    #             max_choice_score = max([choice.scores for choice in question.choices.all()])
    #             total_score += max_choice_score + question.score
    #         elif question.question_type == "checkbox":
    #             choices_total_score = sum([choice.scores for choice in question.choices.all()])
    #             total_score += choices_total_score + question.score
    #         else:
    #             total_score += question.score

    #     _temp = []
    #     for response in responseInfo.response.all():
    #         if response.answer_to.question_type in ["short", "paragraph"]:
    #             if response.answer == response.answer_to.answer_key:
    #                 score += response.answer_to.score
    #         elif response.answer_to.question_type == "multiple choice":
    #             answerKey = None
    #             choice_score = 0
    #             for choice in response.answer_to.choices.all():
    #                 if choice.is_answer:
    #                     answerKey = choice.id
    #                 if choice.id == int(response.answer):
    #                     choice_score = choice.scores
    #             if answerKey is not None and int(answerKey) == int(response.answer):
    #                 score += response.answer_to.score + choice_score
    #         elif response.answer_to.question_type == "checkbox" and response.answer_to.pk not in _temp:
    #             answers = []
    #             answer_keys = []
    #             choice_scores_sum = 0
    #             selected_scores_sum = 0
    #             for resp in responseInfo.response.filter(answer_to__pk=response.answer_to.pk):
    #                 answers.append(int(resp.answer))
    #                 for choice in resp.answer_to.choices.all():
    #                     if choice.is_answer and choice.pk not in answer_keys:
    #                         answer_keys.append(choice.pk)
    #                     if choice.pk == int(resp.answer):
    #                         selected_scores_sum += choice.scores
    #                 _temp.append(response.answer_to.pk)
    #             if set(answers) == set(answer_keys):
    #                 score += response.answer_to.score
    #             score += selected_scores_sum

    return render(request, "index/form/response.html", {
        "form": formInfo,
        "response": responseInfo,
        "score": score,
        "total_score": total_score
    })

def calculate_final_scores(request, code):
    final_scores = {}
    active_forms = Form.objects.filter(is_active=True)
    
    # Получаем параметры фильтрации
    selected_city = request.GET.get('cities')
    selected_gender = request.GET.get('gender')
    age_min = request.GET.get('age_min')
    age_max = request.GET.get('age_max')
    selected_med_region = request.GET.get('med_region')
    selected_med_center = request.GET.get('med_center')
    selected_med_group = request.GET.get('med_group')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    for form in active_forms:
        # Базовый запрос для каждой формы
        responses = Responses.objects.filter(response_to=form)
        
        # Применяем фильтры
        filters = Q()
        if selected_city:
            filters &= Q(responder_city=selected_city)
        if selected_gender:
            filters &= Q(responder_gender=selected_gender)
        if age_min and age_min.strip():
            try:
                filters &= Q(responder_age__gte=int(age_min))
            except (ValueError, TypeError):
                pass
        if age_max and age_max.strip():
            try:
                filters &= Q(responder_age__lte=int(age_max))
            except (ValueError, TypeError):
                pass
        if selected_med_group:
            med_centers = RegionMedCenter.objects.filter(group_id=selected_med_group).values_list('med_center', flat=True)
            filters &= Q(responder_med__in=med_centers)
        if selected_med_region:
            med_centers = RegionMedCenter.objects.filter(region=selected_med_region).values_list('med_center', flat=True)
            filters &= Q(responder_med__in=med_centers)
        if selected_med_center and selected_med_center.strip():
            filters &= Q(responder_med__exact=selected_med_center)
        if date_from and date_from.strip():
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                filters &= Q(createdAt__gte=date_from_obj)
            except (ValueError, TypeError):
                pass
        if date_to and date_to.strip():
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
                filters &= Q(createdAt__lt=date_to_obj)
            except (ValueError, TypeError):
                pass
        
        # Применяем фильтры
        responses = responses.filter(filters)
        
        average_scores = calculate_average_scores(responses, form)
        
        for med_center, data in average_scores.items():
            if med_center not in final_scores:
                final_scores[med_center] = {
                    'forms': {},
                    'total_score': 0,
                    'negative_count': 0,
                    'negative_percentage': 0
                }
            
            if data['total_score'] is not None:
                # Сохраняем процентное значение
                final_scores[med_center]['forms'][form.title] = data['total_score']
    
    # Подсчет негативных ответов и их влияния
    current_form = Form.objects.get(code=code)
    
    # Базовый запрос для текущей формы
    current_responses = Responses.objects.filter(response_to=current_form)
    
    # Применяем фильтры
    filters = Q()
    if selected_city:
        filters &= Q(responder_city=selected_city)
    if selected_gender:
        filters &= Q(responder_gender=selected_gender)
    if age_min and age_min.strip():
        try:
            filters &= Q(responder_age__gte=int(age_min))
        except (ValueError, TypeError):
            pass
    if age_max and age_max.strip():
        try:
            filters &= Q(responder_age__lte=int(age_max))
        except (ValueError, TypeError):
            pass
    if selected_med_group:
        med_centers = RegionMedCenter.objects.filter(group_id=selected_med_group).values_list('med_center', flat=True)
        filters &= Q(responder_med__in=med_centers)
    if selected_med_region:
        med_centers = RegionMedCenter.objects.filter(region=selected_med_region).values_list('med_center', flat=True)
        filters &= Q(responder_med__in=med_centers)
    if selected_med_center and selected_med_center.strip():
        filters &= Q(responder_med__exact=selected_med_center)
    if date_from and date_from.strip():
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            filters &= Q(createdAt__gte=date_from_obj)
        except (ValueError, TypeError):
            pass
    if date_to and date_to.strip():
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            filters &= Q(createdAt__lt=date_to_obj)
        except (ValueError, TypeError):
            pass
    
    # Применяем фильтры
    current_responses = current_responses.filter(filters)
    
    for med_center in final_scores:
        negative_count = 0
        total_forms = 0
        total_percentage = 0
        
        # Подсчет негативных ответов
        med_center_responses = current_responses.filter(responder_med=med_center)
        for response in med_center_responses:
            for answer in response.response.filter(answer_to__is_negative=True, answer='Да'):
                negative_count += 1
        
        # Расчет среднего процента по всем формам
        for score in final_scores[med_center]['forms'].values():
            if score is not None:
                total_percentage += score
                total_forms += 1
        
        final_scores[med_center]['negative_count'] = negative_count
        
        # Расчет влияния негативных ответов
        if total_forms > 0:
            average_percentage = total_percentage / total_forms
            negative_impact = negative_count * 5  # 5% за каждый негативный ответ
            final_scores[med_center]['negative_percentage'] = min(negative_impact, 100)
            final_scores[med_center]['total_score'] = max(0, round(average_percentage * (1 - negative_impact / 100), 2))
        else:
            final_scores[med_center]['total_score'] = 0
            final_scores[med_center]['negative_percentage'] = 0
    
    return final_scores

def export_final_scores(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))

    formInfo = get_object_or_404(Form, code=code)
    if not request.user.is_superuser and not request.user.is_staff:
        return HttpResponseRedirect(reverse("403"))

    # Получаем активные формы и данные
    active_forms = Form.objects.filter(is_active=True)
    final_scores = calculate_final_scores(request, code)

    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Итоговые оценки"

    # Создаем заголовки
    headers = ["Медицинский центр"]
    for form in active_forms:
        headers.append(form.title)
    headers.extend(["Количество жалоб", "Процент влияния жалоб", "Итоговая оценка"])

    # Применяем стили к заголовкам
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        # Устанавливаем ширину столбца
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Заполняем данные
    row = 2
    for med_center, data in final_scores.items():
        col = 1
        # Медцентр
        ws.cell(row=row, column=col, value=med_center)
        col += 1
        
        # Оценки по формам
        for form in active_forms:
            score = data['forms'].get(form.title, 'N/A')
            if score != 'N/A':
                ws.cell(row=row, column=col, value=float(score))
            else:
                ws.cell(row=row, column=col, value=score)
            col += 1
        
        # Количество жалоб
        ws.cell(row=row, column=col, value=data['negative_count'])
        col += 1
        
        # Процент влияния жалоб (делим на 100, так как значение уже в процентах)
        percentage_cell = ws.cell(row=row, column=col, value=data['negative_percentage'] / 100)
        percentage_cell.number_format = '0%'
        col += 1
        
        # Итоговая оценка
        ws.cell(row=row, column=col, value=data['total_score'])
        
        row += 1

    # Создаем HTTP ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Final_Scores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    # Сохраняем файл
    wb.save(response)
    return response

def edit_response(request, code, response_code):
    formInfo = Form.objects.filter(code=code).first()
    if not formInfo:
        return HttpResponseRedirect(reverse('404'))
    
    response = Responses.objects.filter(response_code=response_code, response_to=formInfo).first()
    if not response:
        return HttpResponseRedirect(reverse('404'))
    
    if formInfo.authenticated_responder:
        if not request.user.is_authenticated:
            return HttpResponseRedirect(reverse("login"))
        if response.responder != request.user:
            return HttpResponseRedirect(reverse('403'))
    
    if request.method == "POST":
        if formInfo.authenticated_responder and not response.responder:
            response.responder = request.user
            response.save()
        if formInfo.collect_email:
            response.responder_email = request.POST["email-address"]
            response.save()
        for i in response.response.all():
            i.delete()
        for i in request.POST:
            if i == "csrfmiddlewaretoken" or i == "email-address":
                continue
            question = formInfo.questions.get(id=i)
            for j in request.POST.getlist(i):
                is_skipped = request.POST.get(f'is_skipped_{question.id}') == 'True'
                answer = Answer(answer=j, answer_to=question, is_skipped=is_skipped)
                answer.save()
                response.response.add(answer)
                response.save()
        # if formInfo.is_quiz:
        #     return HttpResponseRedirect(reverse("response", args=[formInfo.code, response.response_code]))
        # else:
        return render(request, "index/form/form_response.html", {
            "form": formInfo,
            "response": response
        })
    return render(request, "index/form/edit_response.html", {
        "form": formInfo,
        "response": response
    })

def delete_responses(request, code):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    formInfo = Form.objects.filter(code = code)
    if formInfo.count() == 0:
        return HttpResponseRedirect(reverse('404'))
    else: formInfo = formInfo[0]
    if not request.user.is_superuser and not request.user.is_staff:
        return HttpResponseRedirect(reverse("403"))
    if request.method == "DELETE":
        responses = Responses.objects.filter(response_to = formInfo)
        for response in responses:
            for i in response.response.all():
                i.delete()
            response.delete()
        return JsonResponse({"message": "Успешно"})
    

@csrf_exempt
def check_question_has_answers(request, question_id):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Forbidden'}, status=403)
    
    from index.models import Questions, Answer
    question = Questions.objects.filter(id=question_id).first()
    
    if not question:
        return JsonResponse({'error': 'Question not found'}, status=404)
    
    has_answers = Answer.objects.filter(answer_to=question).exists()
    
    return JsonResponse({
        'has_answers': has_answers
    })

def load_responses(request, code):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    formInfo = get_object_or_404(Form, code=code)
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 20))

    # Get all responses with prefetch_related
    responses = Responses.objects.filter(response_to=formInfo).select_related(
        'responder'
    ).prefetch_related(
        'response',
        'response__answer_to',
        'response__answer_to__choices'
    ).order_by('-createdAt')

    # Create paginator
    paginator = Paginator(responses, per_page)
    page_obj = paginator.get_page(page)

    # Format response data
    response_data = []
    for response in page_obj:
        data = {
            'id': response.id,
            'username': response.responder_username if response.responder_username else 'Anonymous',
            'url': reverse('response', args=[formInfo.code, response.response_code]),
            'profile_image': None,
            'score': None
        }

        # Add profile image if exists
        if response.responder and hasattr(response.responder, 'images'):
            profile_image = response.responder.images.first()
            if profile_image and profile_image.image:
                data['profile_image'] = profile_image.image.url

        # Add score if form is quiz
        # if formInfo.is_quiz:
        #     data['score'] = f"{calculate_score(response, formInfo)} / {calculate_total_score(formInfo)}"

        response_data.append(data)

    return JsonResponse({
        'responses': response_data,
        'has_next': page_obj.has_next()
    })

def search_responses(request, code):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    formInfo = get_object_or_404(Form, code=code)
    
    # Получаем поисковый запрос, не преобразуя его в нижний регистр,
    # так как Django's icontains сам обрабатывает регистр
    search_term = request.GET.get('term', '')
    
    # Логирование для отладки
    print(f"Search term received: '{search_term}'")
    
    # Get filtered responses - только по имени пользователя
    responses = Responses.objects.filter(
        response_to=formInfo
    ).filter(
        responder_username__icontains=search_term
    ).select_related(
        'responder'
    ).prefetch_related(
        'response',
        'response__answer_to',
        'response__answer_to__choices'
    ).order_by('-createdAt')
    
    # Логирование количества найденных результатов
    print(f"Found {responses.count()} results for '{search_term}'")
    
    # Логирование первых результатов для проверки
    for i, resp in enumerate(responses[:5]):
        print(f"Result {i+1}: '{resp.responder_username}'")

    # Format response data
    response_data = []
    for response in responses:
        data = {
            'id': response.id,
            'username': response.responder_username if response.responder_username else 'Anonymous',
            'url': reverse('response', args=[formInfo.code, response.response_code]),
            'profile_image': None,
            'score': None
        }

        # Add profile image if exists
        if response.responder and hasattr(response.responder, 'images'):
            profile_image = response.responder.images.first()
            if profile_image and profile_image.image:
                data['profile_image'] = profile_image.image.url

        # if formInfo.is_quiz:
        #     data['score'] = f"{calculate_score(response, formInfo)} / {calculate_total_score(formInfo)}"

        response_data.append(data)

    return JsonResponse({'responses': response_data})

def load_table_data(request, code):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    try:
        formInfo = get_object_or_404(Form, code=code)
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))

        # Получаем параметры фильтрации из запроса
        selected_city = request.GET.get('cities')
        selected_gender = request.GET.get('gender')
        age_min = request.GET.get('age_min')
        age_max = request.GET.get('age_max')
        selected_med_region = request.GET.get('med_region')
        selected_med_center = request.GET.get('med_center')
        selected_med_group = request.GET.get('med_group')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        # Базовый запрос
        responses = Responses.objects.filter(response_to=formInfo).select_related(
            'responder'
        ).prefetch_related(
            'response',
            'response__answer_to',
            'response__answer_to__choices'
        ).order_by('-createdAt')
        
        # Применяем фильтры с использованием Q objects для оптимизации
        filters = Q()
        if selected_city:
            filters &= Q(responder_city=selected_city)
        if selected_gender:
            filters &= Q(responder_gender=selected_gender)
        if age_min and age_min.strip():
            try:
                filters &= Q(responder_age__gte=int(age_min))
            except (ValueError, TypeError):
                pass
        if age_max and age_max.strip():
            try:
                filters &= Q(responder_age__lte=int(age_max))
            except (ValueError, TypeError):
                pass
        if selected_med_group:
            med_centers = RegionMedCenter.objects.filter(group_id=selected_med_group).values_list('med_center', flat=True)
            filters &= Q(responder_med__in=med_centers)
        if selected_med_region:
            med_centers = RegionMedCenter.objects.filter(region=selected_med_region).values_list('med_center', flat=True)
            filters &= Q(responder_med__in=med_centers)
        if selected_med_center and selected_med_center.strip():
            filters &= Q(responder_med__exact=selected_med_center)
        if date_from and date_from.strip():
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                filters &= Q(createdAt__gte=date_from_obj)
            except (ValueError, TypeError):
                pass
        if date_to and date_to.strip():
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
                filters &= Q(createdAt__lt=date_to_obj)
            except (ValueError, TypeError):
                pass
        
        # Применяем фильтры по выбранным вариантам ответов
        for key, value in request.GET.items():
            if key.startswith('question-') and value:
                question_id = key.replace('question-', '')
                try:
                    question = Questions.objects.get(id=question_id)
                    if question.question_type in ["multiple choice", "checkbox"]:
                        choice_ids = request.GET.getlist(key)
                        if choice_ids:
                            answers = Answer.objects.filter(
                                answer_to=question,
                                answer__in=choice_ids
                            )
                            response_ids = answers.values_list('response_id', flat=True)
                            filters &= Q(id__in=response_ids)
                except Questions.DoesNotExist:
                    pass

        responses = responses.filter(filters)

        # Create paginator
        paginator = Paginator(responses, per_page)
        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            return JsonResponse({'rows': [], 'has_next': False})

        # Format response data
        rows = []
        for response in page_obj:
            # Render row HTML using a template
            row_html = render_to_string('index/tables/response_table_row.html', {
                'response': response,
                'form': formInfo,
                'request': request
            })
            rows.append({'html': row_html})

        return JsonResponse({
            'rows': rows,
            'has_next': page_obj.has_next()
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def load_med_centers_data(request, code):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    try:
        formInfo = get_object_or_404(Form, code=code)
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))

        # Получаем параметры фильтрации из запроса
        selected_city = request.GET.get('cities')
        selected_gender = request.GET.get('gender')
        age_min = request.GET.get('age_min')
        age_max = request.GET.get('age_max')
        selected_med_region = request.GET.get('med_region')
        selected_med_center = request.GET.get('med_center')
        selected_med_group = request.GET.get('med_group')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        # Базовый запрос
        responses = Responses.objects.filter(response_to=formInfo)
        
        # Применяем фильтры с использованием Q objects для оптимизации
        filters = Q()
        if selected_city:
            filters &= Q(responder_city=selected_city)
        if selected_gender:
            filters &= Q(responder_gender=selected_gender)
        if age_min and age_min.strip():
            try:
                filters &= Q(responder_age__gte=int(age_min))
            except (ValueError, TypeError):
                pass
        if age_max and age_max.strip():
            try:
                filters &= Q(responder_age__lte=int(age_max))
            except (ValueError, TypeError):
                pass
        if selected_med_group:
            med_centers = RegionMedCenter.objects.filter(group_id=selected_med_group).values_list('med_center', flat=True)
            filters &= Q(responder_med__in=med_centers)
        if selected_med_region:
            med_centers = RegionMedCenter.objects.filter(region=selected_med_region).values_list('med_center', flat=True)
            filters &= Q(responder_med__in=med_centers)
        if selected_med_center and selected_med_center.strip():
            filters &= Q(responder_med__exact=selected_med_center)
        if date_from and date_from.strip():
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                filters &= Q(createdAt__gte=date_from_obj)
            except (ValueError, TypeError):
                pass
        if date_to and date_to.strip():
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
                filters &= Q(createdAt__lt=date_to_obj)
            except (ValueError, TypeError):
                pass
        
        # Применяем фильтры по выбранным вариантам ответов
        for key, value in request.GET.items():
            if key.startswith('question-') and value:
                question_id = key.replace('question-', '')
                try:
                    question = Questions.objects.get(id=question_id)
                    if question.question_type in ["multiple choice", "checkbox"]:
                        choice_ids = request.GET.getlist(key)
                        if choice_ids:
                            answers = Answer.objects.filter(
                                answer_to=question,
                                answer__in=choice_ids
                            )
                            response_ids = answers.values_list('response_id', flat=True)
                            filters &= Q(id__in=response_ids)
                except Questions.DoesNotExist:
                    pass

        responses = responses.filter(filters)
        
        # Calculate average scores
        average_scores = calculate_average_scores(responses, formInfo)
        med_center_stats = get_med_center_stats(formInfo, responses)
        
        average_scores_sorted = sorted(
            [item for item in average_scores.items() if item[1]['total_score'] is not None],
            key=lambda x: x[1]['total_score'],
            reverse=True
        )

        # Create paginator
        paginator = Paginator(average_scores_sorted, per_page)
        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            return JsonResponse({'rows': [], 'has_next': False})

        # Format response data
        rows = []
        for med_center, scores in page_obj:
            # Render row HTML using a template
            row_html = render_to_string('index/tables/med_center_table_row.html', {
                'med_center': med_center,
                'scores': scores,
                'form': formInfo,
                'med_center_stats': med_center_stats,
                'request': request
            })
            rows.append({'html': row_html})

        return JsonResponse({
            'rows': rows,
            'has_next': page_obj.has_next()
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def load_total_scores_data(request, code):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    try:
        formInfo = get_object_or_404(Form, code=code)
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))

        # Получаем параметры фильтрации из запроса
        selected_city = request.GET.get('cities')
        selected_gender = request.GET.get('gender')
        age_min = request.GET.get('age_min')
        age_max = request.GET.get('age_max')
        selected_med_region = request.GET.get('med_region')
        selected_med_center = request.GET.get('med_center')
        selected_med_group = request.GET.get('med_group')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        # Сохраняем параметры фильтрации во временном словаре для calculate_final_scores
        filter_params = {
            'cities': selected_city,
            'gender': selected_gender,
            'age_min': age_min,
            'age_max': age_max,
            'med_region': selected_med_region,
            'med_center': selected_med_center,
            'med_group': selected_med_group,
            'date_from': date_from,
            'date_to': date_to
        }
        
        # Сохраняем параметры в request для использования в calculate_final_scores
        request.GET = request.GET.copy()
        for key, value in filter_params.items():
            if value:
                request.GET[key] = value

        # Calculate final scores
        final_scores = calculate_final_scores(request, code)
        
        # Фильтрация по медцентрам
        if selected_med_center:
            final_scores = {k: v for k, v in final_scores.items() if k == selected_med_center}
        if selected_med_region:
            med_centers = RegionMedCenter.objects.filter(region=selected_med_region).values_list('med_center', flat=True)
            final_scores = {k: v for k, v in final_scores.items() if k in med_centers}
        if selected_med_group:
            med_centers = RegionMedCenter.objects.filter(group_id=selected_med_group).values_list('med_center', flat=True)
            final_scores = {k: v for k, v in final_scores.items() if k in med_centers}
        
        final_scores_list = sorted(
            final_scores.items(),
            key=lambda x: x[1]['total_score'] if x[1]['total_score'] is not None else -1,
            reverse=True
        )

        # Create paginator
        paginator = Paginator(final_scores_list, per_page)
        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            return JsonResponse({'rows': [], 'has_next': False})

        # Format response data
        rows = []
        for med_center, data in page_obj:
            # Render row HTML using a template
            row_html = render_to_string('index/tables/total_scores_table_row.html', {
                'med_center': med_center,
                'data': data,
                'active_forms': Form.objects.filter(is_active=True),
                'request': request
            })
            rows.append({'html': row_html})

        return JsonResponse({
            'rows': rows,
            'has_next': page_obj.has_next()
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def load_final_scores_chart_data(request, code):
    """Load data for final scores chart asynchronously"""
    try:
        formInfo = Form.objects.get(code=code)
        if not formInfo.creator == request.user and not request.user.is_superuser:
            return JsonResponse({'error': 'Unauthorized'}, status=403)

        final_scores = calculate_final_scores(request, code)
        active_forms = Form.objects.filter(is_active=True)

        chart_data = {
            'labels': list(final_scores.keys()),
            'datasets': []
        }

        # Generate colors for each dataset
        colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40']

        for idx, form in enumerate(active_forms):
            color = colors[idx % len(colors)]  # Cycle through colors if more forms than colors
            dataset = {
                'label': form.title,
                'data': [final_scores[med_center]['forms'].get(form.title, 0) for med_center in final_scores.keys()],
                'backgroundColor': color,
                'borderColor': color,
                'borderWidth': 2,
                'fill': False
            }
            chart_data['datasets'].append(dataset)

        return JsonResponse(chart_data)
    except Form.DoesNotExist:
        return JsonResponse({'error': 'Form not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def load_negative_impact_chart_data(request, code):
    """Load data for negative impact chart asynchronously"""
    try:
        formInfo = Form.objects.get(code=code)
        if not formInfo.creator == request.user and not request.user.is_superuser:
            return JsonResponse({'error': 'Unauthorized'}, status=403)

        final_scores = calculate_final_scores(request, code)

        chart_data = {
            'labels': list(final_scores.keys()),
            'datasets': [
                {
                    'label': 'Итоговая оценка',
                    'data': [data['total_score'] for data in final_scores.values()],
                    'backgroundColor': '#4CAF50',
                    'borderColor': '#4CAF50',
                    'borderWidth': 2,
                    'yAxisID': 'y-axis-1',
                    'order': 2
                },
                {
                    'label': 'Количество жалоб',
                    'data': [data['negative_count'] for data in final_scores.values()],
                    'backgroundColor': '#FF3030',
                    'borderColor': '#FF3030',
                    'borderWidth': 2,
                    'type': 'line',
                    'yAxisID': 'y-axis-2',
                    'order': 1
                }
            ]
        }

        return JsonResponse(chart_data)
    except Form.DoesNotExist:
        return JsonResponse({'error': 'Form not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_cache_key(request, formInfo, data_type):
    """Генерация ключа кэша на основе параметров запроса"""
    params = request.GET.copy()
    params['form_code'] = formInfo.code
    params['data_type'] = data_type
    return f"analytics_data:{urlquote(str(sorted(params.items())))}"

@login_required
def load_analytics_data(request, code):
    """Загрузка аналитических данных для формы с кэшированием"""
    formInfo = get_object_or_404(Form, code=code)
    data_type = request.GET.get('data_type')
    
    # Генерируем ключ кэша
    cache_key = get_cache_key(request, formInfo, data_type)
    
    # Пробуем получить данные из кэша
    cached_data = cache.get(cache_key)
    if cached_data:
        return JsonResponse(cached_data)
    
    # Если данных в кэше нет, вычисляем их
    all_responses = get_filtered_responses_queryset(request, formInfo)
    
    if data_type == 'summary':
        questions = formInfo.questions.prefetch_related('choices').exclude(question_type="title")
        responsesSummary = []
        choiceAnswered = {}
        choices_dict = {}
        
        for question in questions:
            answers = Answer.objects.filter(
                answer_to=question.id, 
                is_skipped=False,
                response__in=all_responses
            ).select_related('answer_to').prefetch_related('answer_to__choices')
            
            if question.question_type in ["multiple choice", "checkbox"]:
                choiceAnswered[question.id] = {}
                for answer in answers:
                    try:
                        choice = answer.answer_to.choices.get(id=answer.answer)
                        choices_dict[answer.answer] = choice.choice
                        unique_choice_key = f"{choice.choice}"
                        choiceAnswered[question.id][unique_choice_key] = choiceAnswered[question.id].get(unique_choice_key, 0) + 1
                    except Choices.DoesNotExist:
                        continue
            
            responsesSummary.append({"question": question, "answers": answers})
            
        data = {
            'responsesSummary': responsesSummary,
            'choiceAnswered': choiceAnswered,
            'choices_dict': choices_dict,
        }
        
    elif data_type == 'averages':
        average_scores = calculate_average_scores(all_responses, formInfo)
        average_scores_sorted = sorted(
            [item for item in average_scores.items() if item[1]['total_score'] is not None],
            key=lambda x: x[1]['total_score'],
            reverse=True
        )
        data = {
            'average_scores': average_scores,
            'average_scores_sorted': average_scores_sorted,
        }
        
    elif data_type == 'med_stats':
        data = {
            'med_center_stats': get_med_center_stats(formInfo, all_responses),
        }
        
    elif data_type == 'range_slider':
        # For range slider data, we want to show the full year of data
        # but only for medical centers that match the other filters
        # We need to create a new queryset that ignores date filters
        
        # Get all filter parameters except dates
        selected_city = request.GET.get('cities')
        selected_gender = request.GET.get('gender')
        age_min = request.GET.get('age_min')
        age_max = request.GET.get('age_max')
        selected_med_region = request.GET.get('med_region')
        selected_med_center = request.GET.get('med_center')
        selected_med_group = request.GET.get('med_group')
        
        # Create a new queryset without date filters
        responses_without_date_filters = get_filtered_responses(
            formInfo, 
            age_min, 
            age_max, 
            selected_gender, 
            [selected_city] if selected_city else None,
            selected_med_center
        )
        
        # Apply med_center filters if needed
        if selected_med_group:
            med_centers = RegionMedCenter.objects.filter(group_id=selected_med_group).values_list('med_center', flat=True)
            responses_without_date_filters = responses_without_date_filters.filter(responder_med__in=med_centers)
        
        if selected_med_region:
            med_centers = RegionMedCenter.objects.filter(region=selected_med_region).values_list('med_center', flat=True)
            responses_without_date_filters = responses_without_date_filters.filter(responder_med__in=med_centers)
        
        data = {
            'range_slider_data': get_range_slider_data(formInfo, responses_without_date_filters),
        }
        
    elif data_type == 'final_scores':
        data = {
            'final_scores': calculate_final_scores(request, code),
        }
        
    else:
        return JsonResponse({'error': 'Invalid data type'}, status=400)
    
    # Сохраняем результат в кэш на 5 минут
    cache.set(cache_key, data, 300)
    
    return JsonResponse(data)

def get_filtered_responses_queryset(request, formInfo):
    """Получение отфильтрованного queryset ответов"""
    selected_city = request.GET.get('cities')
    selected_gender = request.GET.get('gender')
    age_min = request.GET.get('age_min')
    age_max = request.GET.get('age_max')
    selected_med_region = request.GET.get('med_region')
    selected_med_center = request.GET.get('med_center')
    selected_med_group = request.GET.get('med_group')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    all_responses = Responses.objects.filter(response_to=formInfo).select_related(
        'responder'
    ).prefetch_related(
        'response',
        'response__answer_to',
        'response__answer_to__choices'
    )
    
    filters = Q()
    if selected_city:
        filters &= Q(responder_city=selected_city)
    if selected_gender:
        filters &= Q(responder_gender=selected_gender)
    if age_min and age_min.strip():
        try:
            filters &= Q(responder_age__gte=int(age_min))
        except (ValueError, TypeError):
            pass
    if age_max and age_max.strip():
        try:
            filters &= Q(responder_age__lte=int(age_max))
        except (ValueError, TypeError):
            pass
    if selected_med_group:
        med_centers = RegionMedCenter.objects.filter(group_id=selected_med_group).values_list('med_center', flat=True)
        filters &= Q(responder_med__in=med_centers)
    if selected_med_region:
        med_centers = RegionMedCenter.objects.filter(region=selected_med_region).values_list('med_center', flat=True)
        filters &= Q(responder_med__in=med_centers)
    if selected_med_center and selected_med_center.strip():
        # Ensure exact match for med_center
        filters &= Q(responder_med__exact=selected_med_center)
    if date_from and date_from.strip():
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            filters &= Q(createdAt__gte=date_from_obj)
        except (ValueError, TypeError):
            pass
    if date_to and date_to.strip():
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            filters &= Q(createdAt__lt=date_to_obj)
        except (ValueError, TypeError):
            pass
        
    return all_responses.filter(filters)